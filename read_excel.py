import openpyxl, json

wb = openpyxl.load_workbook('lmt-day-booth/game-content.xlsx', data_only=True)
output = {}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        rows.append([str(c) if c is not None else '' for c in row])
    output[sheet_name] = rows

for name, rows in output.items():
    print(f"\n=== SHEET: {name} ({len(rows)} rows) ===")
    for i, row in enumerate(rows[:3]):
        print(f"  Row {i}: {row[:6]}")
    if len(rows) > 3:
        print(f"  ... ({len(rows)-3} more rows)")

print("\n\nFULL JSON:")
print(json.dumps(output, ensure_ascii=False, indent=2))
